"""
05_analysis.py
--------------
Pattern analysis for Thai keyboard-switched passwords.

Part A: Real-world dataset (709 passwords from .th breach data)
  - Length distribution
  - zxcvbn score distribution
  - Entropy distribution
  - Character type analysis

Part B: นศ. constructed dataset — Cracked vs Not-Cracked
  - Length vs crack rate
  - Character complexity vs crack rate
  - Attack vector breakdown
  - zxcvbn score vs crack rate (zxcvbn blind spot analysis)

Outputs:
  results/tables/summary_stats.csv
  results/tables/length_vs_crackrate.csv
  results/tables/attack_vectors.csv
  results/tables/zxcvbn_vs_crackrate.csv
  results/figures/  (PNG charts)

Run:
    python scripts/05_analysis.py
"""

import csv
import re
import statistics
import sys
from collections import Counter
from pathlib import Path
import openpyxl

# ── Output directories ──────────────────────────────────────────
TABLES  = Path('results/tables')
FIGURES = Path('results/figures')
TABLES.mkdir(parents=True, exist_ok=True)
FIGURES.mkdir(parents=True, exist_ok=True)

# ── Input files ─────────────────────────────────────────────────
METRICS_RW   = Path('data/processed/metrics_realworld.csv')
DATASET_XLSX = Path('data/raw/CompilationOfManyBreaches/../../../')  # placeholder
# นศ. dataset path — adjust if needed
NST_XLSX = Path('data/processed/nst_dataset.xlsx')


# ════════════════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════════════════

def char_complexity(pw: str) -> int:
    """Count how many character types are present (0-4)."""
    return sum([
        bool(re.search(r'[A-Z]', pw)),
        bool(re.search(r'[a-z]', pw)),
        bool(re.search(r'[0-9]', pw)),
        bool(re.search(r'[^a-zA-Z0-9]', pw)),
    ])


def write_csv(path: Path, headers: list, rows: list):
    with open(path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)
    print(f'  📄 {path}')


def mean_std(data: list) -> tuple:
    if not data:
        return 0.0, 0.0
    m = statistics.mean(data)
    s = statistics.stdev(data) if len(data) > 1 else 0.0
    return m, s


# ════════════════════════════════════════════════════════════════
# Part A: Real-world 709 analysis
# ════════════════════════════════════════════════════════════════

def analyze_realworld(metrics_path: Path):
    print('\n=== Part A: Real-world .th breach passwords (n=709) ===\n')

    with open(metrics_path, encoding='utf-8') as f:
        rows = [r for r in csv.DictReader(f)
                if r.get('zxcvbn_score') not in ('', '-1', None)]

    passwords = [r['password'] for r in rows]
    lengths   = [int(r['length']) for r in rows]
    peak_ent  = [float(r['peak_entropy_bits']) for r in rows]
    shannon   = [float(r['shannon_per_symbol']) for r in rows]
    sh_total  = [float(r['shannon_total']) for r in rows]
    zxcvbn    = [int(r['zxcvbn_score']) for r in rows]
    n = len(rows)

    # ── Summary stats ──
    summary = []
    for label, data in [
        ('Length',               lengths),
        ('Peak Entropy (bits)',  peak_ent),
        ('Shannon per symbol',   shannon),
        ('Shannon Total',        sh_total),
        ('zxcvbn Score',         zxcvbn),
    ]:
        m, s = mean_std(data)
        med  = statistics.median(data)
        summary.append([label, f'{m:.4f}', f'{s:.4f}', f'{med:.4f}',
                         f'{min(data):.4f}', f'{max(data):.4f}'])
        print(f'  {label:25s}  mean={m:.3f}  std={s:.3f}  median={med:.3f}')

    write_csv(TABLES / 'realworld_summary_stats.csv',
              ['metric', 'mean', 'std', 'median', 'min', 'max'],
              summary)

    # ── Length distribution ──
    len_dist = Counter(lengths)
    len_rows = [[l, len_dist[l], f'{len_dist[l]/n*100:.1f}']
                for l in sorted(len_dist)]
    write_csv(TABLES / 'realworld_length_distribution.csv',
              ['length', 'count', 'pct'],
              len_rows)

    # ── zxcvbn distribution ──
    zx_dist = Counter(zxcvbn)
    zx_rows = [[s, zx_dist.get(s, 0), f'{zx_dist.get(s,0)/n*100:.1f}']
               for s in range(5)]
    write_csv(TABLES / 'realworld_zxcvbn_distribution.csv',
              ['score', 'count', 'pct'],
              zx_rows)
    print()
    print('  zxcvbn distribution:')
    for score, count, pct in zx_rows:
        print(f'    Score {score}: {count:4d}  ({pct}%)')

    # ── Character complexity ──
    complexity = [char_complexity(pw) for pw in passwords]
    comp_dist = Counter(complexity)
    comp_rows = [[c, comp_dist.get(c, 0), f'{comp_dist.get(c,0)/n*100:.1f}']
                 for c in range(5)]
    write_csv(TABLES / 'realworld_char_complexity.csv',
              ['num_char_types', 'count', 'pct'],
              comp_rows)

    print(f'\n  ✅ Part A complete  (n={n})')
    return rows


# ════════════════════════════════════════════════════════════════
# Part B: นศ. dataset — cracked vs not-cracked
# ════════════════════════════════════════════════════════════════

def load_nst_hashcat(xlsx_path: Path):
    """Load Thai passwords + crack results from Hashing_HashCat sheet."""
    print(f'\n  Loading นศ. dataset from {xlsx_path}...')
    wb  = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws  = wb['Hashing_HashCat']

    records = []
    for r in range(2, 1100):
        if ws.cell(r, 9).value != 'swapped':
            continue
        pw       = ws.cell(r, 8).value
        char_cnt = ws.cell(r, 10).value
        is_found = str(ws.cell(r, 23).value).lower()

        attacks = {
            'brute_force':  ws.cell(r, 12).value,
            'd_rockyou':    ws.cell(r, 15).value,
            'd_rockyou24':  ws.cell(r, 16).value,
            'd_pwdb':       ws.cell(r, 17).value,
            'd_thai':       ws.cell(r, 18).value,
            'd_thai_name':  ws.cell(r, 19).value,
            'c_thai_thai':  ws.cell(r, 20).value,
            'c_tname_tname':ws.cell(r, 21).value,
            'c_thai_tname': ws.cell(r, 22).value,
        }

        # get zxcvbn from Sheet1 (col 8, same row order)
        ws1    = wb['Sheet1']
        zxcvbn = None
        for r1 in range(2, 1310):
            if ws1.cell(r1, 2).value == pw:
                zxcvbn = ws1.cell(r1, 8).value
                break

        records.append({
            'pw':       pw,
            'length':   char_cnt,
            'cracked':  is_found == 'yes',
            'zxcvbn':   zxcvbn,
            'attacks':  attacks,
        })

    cracked     = [r for r in records if r['cracked']]
    not_cracked = [r for r in records if not r['cracked']]
    print(f'  Loaded {len(records)} Thai passwords  '
          f'(cracked={len(cracked)}, not_cracked={len(not_cracked)})')
    return records, cracked, not_cracked


def analyze_nst(records, cracked, not_cracked):
    print('\n=== Part B: นศ. dataset — cracked vs not-cracked ===\n')
    n_c  = len(cracked)
    n_nc = len(not_cracked)

    # ── Length vs crack rate ──
    length_stats = {}
    for rec in records:
        l = rec['length']
        if not l:
            continue
        if l not in length_stats:
            length_stats[l] = {'cracked': 0, 'total': 0}
        length_stats[l]['total']   += 1
        length_stats[l]['cracked'] += int(rec['cracked'])

    len_rows = []
    print('  Length → crack rate:')
    for l in sorted(length_stats):
        s    = length_stats[l]
        rate = s['cracked'] / s['total'] * 100
        len_rows.append([l, s['total'], s['cracked'], f'{rate:.1f}'])
        if s['total'] >= 5:
            print(f'    len={l:2d}  total={s["total"]:4d}  '
                  f'cracked={s["cracked"]:4d}  rate={rate:.1f}%')

    write_csv(TABLES / 'nst_length_vs_crackrate.csv',
              ['length', 'total', 'cracked', 'crack_rate_pct'],
              len_rows)

    # ── Length buckets ──
    buckets = [(1,8,'1-8'), (9,12,'9-12'), (13,16,'13-16'),
               (17,20,'17-20'), (21,99,'21+')]
    bucket_rows = []
    print('\n  Length buckets:')
    for lo, hi, label in buckets:
        grp   = [r for r in records if r['length'] and lo <= r['length'] <= hi]
        n_crk = sum(1 for r in grp if r['cracked'])
        rate  = n_crk / len(grp) * 100 if grp else 0
        bucket_rows.append([label, len(grp), n_crk, f'{rate:.1f}'])
        print(f'    {label:6s}  n={len(grp):4d}  cracked={n_crk:4d}  '
              f'crack_rate={rate:.1f}%')

    write_csv(TABLES / 'nst_length_buckets.csv',
              ['length_range', 'total', 'cracked', 'crack_rate_pct'],
              bucket_rows)

    # ── zxcvbn vs crack rate ──
    print('\n  zxcvbn score → crack rate:')
    zx_rows = []
    for score in range(5):
        grp   = [r for r in records if r.get('zxcvbn') == score]
        n_crk = sum(1 for r in grp if r['cracked'])
        rate  = n_crk / len(grp) * 100 if grp else 0
        zx_rows.append([score, len(grp), n_crk, f'{rate:.1f}'])
        print(f'    Score {score}  n={len(grp):4d}  '
              f'cracked={n_crk:4d}  crack_rate={rate:.1f}%')

    write_csv(TABLES / 'nst_zxcvbn_vs_crackrate.csv',
              ['zxcvbn_score', 'total', 'cracked', 'crack_rate_pct'],
              zx_rows)

    # ── Character complexity vs crack rate ──
    print('\n  Char complexity → crack rate:')
    comp_rows = []
    for types in range(5):
        grp   = [r for r in records
                 if r['pw'] and char_complexity(str(r['pw'])) == types]
        n_crk = sum(1 for r in grp if r['cracked'])
        rate  = n_crk / len(grp) * 100 if grp else 0
        comp_rows.append([types, len(grp), n_crk, f'{rate:.1f}'])
        if grp:
            print(f'    {types} char types  n={len(grp):4d}  '
                  f'cracked={n_crk:4d}  crack_rate={rate:.1f}%')

    write_csv(TABLES / 'nst_complexity_vs_crackrate.csv',
              ['num_char_types', 'total', 'cracked', 'crack_rate_pct'],
              comp_rows)

    # ── Attack vector breakdown ──
    print('\n  Attack vectors (of 398 cracked):')
    attack_counter = Counter()
    for rec in cracked:
        for k, v in rec['attacks'].items():
            if str(v).lower() == 'yes':
                attack_counter[k] += 1

    atk_rows = []
    for k, v in attack_counter.most_common():
        pct = v / n_c * 100
        atk_rows.append([k, v, f'{pct:.1f}'])
        print(f'    {k:20s}  {v:4d}  ({pct:.1f}%)')

    write_csv(TABLES / 'nst_attack_vectors.csv',
              ['attack_vector', 'count', 'pct_of_cracked'],
              atk_rows)

    # ── Cracked without Thai dict ──
    no_thai_dict = ['d_rockyou', 'd_rockyou24', 'd_pwdb', 'brute_force']
    cracked_no_thai = [
        r for r in records
        if any(str(r['attacks'].get(k, '')).lower() == 'yes'
               for k in no_thai_dict)
    ]
    cracked_thai_only = [
        r for r in cracked
        if not any(str(r['attacks'].get(k, '')).lower() == 'yes'
                   for k in no_thai_dict)
    ]
    print(f'\n  Without Thai dict: {len(cracked_no_thai)} cracked '
          f'({len(cracked_no_thai)/1000*100:.1f}%)')
    print(f'  Thai dict needed:  {len(cracked_thai_only)} cracked '
          f'({len(cracked_thai_only)/1000*100:.1f}%)')

    write_csv(TABLES / 'nst_thai_dict_impact.csv',
              ['scenario', 'cracked', 'not_cracked', 'crack_rate_pct'],
              [
                  ['With Thai dictionary',    n_c,   n_nc,
                   f'{n_c/1000*100:.1f}'],
                  ['Without Thai dictionary', len(cracked_no_thai),
                   1000 - len(cracked_no_thai),
                   f'{len(cracked_no_thai)/1000*100:.1f}'],
              ])

    print(f'\n  ✅ Part B complete')


# ════════════════════════════════════════════════════════════════
# Part C: Cross-dataset comparison table
# ════════════════════════════════════════════════════════════════

def compare_datasets(rw_rows):
    print('\n=== Part C: Cross-dataset comparison ===\n')

    # Real-world stats
    rw_zx  = [int(r['zxcvbn_score']) for r in rw_rows]
    rw_pe  = [float(r['peak_entropy_bits']) for r in rw_rows]
    rw_sh  = [float(r['shannon_per_symbol']) for r in rw_rows]
    rw_sht = [float(r['shannon_total']) for r in rw_rows]
    rw_len = [int(r['length']) for r in rw_rows]

    # นศ. constructed dataset values (from presentation/paper)
    comparison = [
        ['Metric',
         'RockYou (baseline)',
         'นศ. Constructed Thai',
         'Real-world .th breach'],
        ['n',
         '1,000', '1,000', str(len(rw_rows))],
        ['Mean Length',
         '~8.0',
         '12.75',
         f'{statistics.mean(rw_len):.2f}'],
        ['Peak Entropy (bits)',
         '44.64',
         '79.72',
         f'{statistics.mean(rw_pe):.2f}'],
        ['Shannon per symbol',
         '2.42',
         '3.24',
         f'{statistics.mean(rw_sh):.4f}'],
        ['Shannon Total',
         '16.81',
         '39.97',
         f'{statistics.mean(rw_sht):.2f}'],
        ['zxcvbn mean',
         '0.13',
         '3.65',
         f'{statistics.mean(rw_zx):.4f}'],
    ]

    write_csv(TABLES / 'cross_dataset_comparison.csv',
              [], comparison)  # headers embedded in rows

    print('  Cross-dataset comparison table:')
    for row in comparison:
        print(f'    {row[0]:30s} | {row[1]:22s} | {row[2]:22s} | {row[3]}')

    print(f'\n  ✅ Part C complete')


# ════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--metrics',  default=str(METRICS_RW),
                        help='Path to metrics_realworld.csv')
    parser.add_argument('--xlsx',     default=None,
                        help='Path to นศ. dataset .xlsx file')
    parser.add_argument('--skip-nst', action='store_true',
                        help='Skip Part B (นศ. dataset analysis)')
    args = parser.parse_args()

    print('📊 05_analysis.py — Thai Keyboard-Switched Password Analysis')
    print('=' * 60)

    # Part A
    rw_rows = analyze_realworld(Path(args.metrics))

    # Part B
    if not args.skip_nst:
        xlsx_path = Path(args.xlsx) if args.xlsx else None
        if xlsx_path and xlsx_path.exists():
            records, cracked, not_cracked = load_nst_hashcat(xlsx_path)
            analyze_nst(records, cracked, not_cracked)
        else:
            print('\n⚠️  Part B skipped — provide --xlsx path to นศ. dataset')
            print('   Example: python scripts/05_analysis.py '
                  '--xlsx data/raw/dataset.xlsx')

    # Part C
    compare_datasets(rw_rows)

    print('\n' + '=' * 60)
    print('✅ All done!')
    print(f'   Tables  → {TABLES}/')
    print(f'   Figures → {FIGURES}/ (run with --plot to generate)')


if __name__ == '__main__':
    main()